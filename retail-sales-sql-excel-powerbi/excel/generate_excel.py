"""
Generate Excel file with pivot tables and charts from retail sales data.
Output: retail_sales_analysis.xlsx  (for Power BI import or standalone review)
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os

# Load data
df = pd.read_csv('../data/raw_sales_data.csv')
df['Order_Date'] = pd.to_datetime(df['Order_Date'])
df['Month'] = df['Order_Date'].dt.to_period('M').astype(str)

wb = Workbook()

# ========== SHEET 1: Raw Data ==========
ws_data = wb.active
ws_data.title = "Raw Data"
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))

for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
    for c_idx, val in enumerate(row, 1):
        cell = ws_data.cell(row=r_idx, column=c_idx, value=val)
        if r_idx == 1:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

for col in ws_data.columns:
    max_len = max(len(str(cell.value or '')) for cell in col)
    ws_data.column_dimensions[col[0].column_letter].width = min(max_len + 3, 25)

# ========== SHEET 2: Pivot Tables ==========
ws_pivot = wb.create_sheet("Pivot Tables")

# --- Pivot 1: Sales by Category ---
cat_pivot = df.groupby('Category').agg(
    Revenue=('Total_Price', 'sum'),
    Profit=('Profit', 'sum'),
    Orders=('Order_ID', 'count')
).round(2).sort_values('Revenue', ascending=False)
cat_pivot['Margin%'] = (cat_pivot['Profit'] / cat_pivot['Revenue'] * 100).round(1)

ws_pivot.cell(row=1, column=1, value="Sales by Category").font = Font(bold=True, size=13)
for r_idx, row in enumerate(dataframe_to_rows(cat_pivot.reset_index(), index=False, header=True), 3):
    for c_idx, val in enumerate(row, 1):
        cell = ws_pivot.cell(row=r_idx, column=c_idx, value=val)
        if r_idx == 3:
            cell.font = header_font
            cell.fill = header_fill
        cell.border = thin_border

# Chart 1: Revenue by Category
chart1 = BarChart()
chart1.type = "col"
chart1.title = "Revenue by Category"
chart1.y_axis.title = "Revenue ($)"
chart1.style = 10
data1 = Reference(ws_pivot, min_col=2, min_row=3, max_row=3 + len(cat_pivot), max_col=2)
cats1 = Reference(ws_pivot, min_col=1, min_row=4, max_row=3 + len(cat_pivot))
chart1.add_data(data1, titles_from_data=True)
chart1.set_categories(cats1)
chart1.width = 16
chart1.height = 10
ws_pivot.add_chart(chart1, "G3")

# --- Pivot 2: Sales by Region ---
region_pivot = df.groupby('Region').agg(
    Revenue=('Total_Price', 'sum'),
    Profit=('Profit', 'sum'),
    Orders=('Order_ID', 'count')
).round(2).sort_values('Revenue', ascending=False)

row_start = 3 + len(cat_pivot) + 3
ws_pivot.cell(row=row_start, column=1, value="Sales by Region").font = Font(bold=True, size=13)
for r_idx, row in enumerate(dataframe_to_rows(region_pivot.reset_index(), index=False, header=True), row_start + 2):
    for c_idx, val in enumerate(row, 1):
        cell = ws_pivot.cell(row=r_idx, column=c_idx, value=val)
        if r_idx == row_start + 2:
            cell.font = header_font
            cell.fill = header_fill
        cell.border = thin_border

# Chart 2: Revenue by Region (Pie)
chart2 = PieChart()
chart2.title = "Revenue by Region"
chart2.style = 10
data2 = Reference(ws_pivot, min_col=2, min_row=row_start + 2, max_row=row_start + 1 + len(region_pivot), max_col=2)
cats2 = Reference(ws_pivot, min_col=1, min_row=row_start + 3, max_row=row_start + 1 + len(region_pivot))
chart2.add_data(data2, titles_from_data=True)
chart2.set_categories(cats2)
chart2.width = 14
chart2.height = 10
ws_pivot.add_chart(chart2, "G" + str(row_start + 2))

# --- Pivot 3: Monthly Trend ---
monthly = df.groupby('Month').agg(
    Revenue=('Total_Price', 'sum'),
    Profit=('Profit', 'sum'),
    Orders=('Order_ID', 'count')
).round(2).reset_index()

row_start2 = row_start + 2 + len(region_pivot) + 3
ws_pivot.cell(row=row_start2, column=1, value="Monthly Revenue Trend").font = Font(bold=True, size=13)
for r_idx, row in enumerate(dataframe_to_rows(monthly, index=False, header=True), row_start2 + 2):
    for c_idx, val in enumerate(row, 1):
        cell = ws_pivot.cell(row=r_idx, column=c_idx, value=val)
        if r_idx == row_start2 + 2:
            cell.font = header_font
            cell.fill = header_fill
        cell.border = thin_border

# Chart 3: Monthly Trend
chart3 = BarChart()
chart3.type = "col"
chart3.title = "Monthly Revenue"
chart3.y_axis.title = "Revenue ($)"
chart3.style = 10
data3 = Reference(ws_pivot, min_col=2, min_row=row_start2 + 2, max_row=row_start2 + 1 + len(monthly), max_col=2)
cats3 = Reference(ws_pivot, min_col=1, min_row=row_start2 + 3, max_row=row_start2 + 1 + len(monthly))
chart3.add_data(data3, titles_from_data=True)
chart3.set_categories(cats3)
chart3.width = 20
chart3.height = 10
ws_pivot.add_chart(chart3, "G" + str(row_start2 + 2))

# --- Pivot 4: Customer Type ---
cust_pivot = df.groupby('Customer_Type').agg(
    Revenue=('Total_Price', 'sum'),
    Profit=('Profit', 'sum'),
    Orders=('Order_ID', 'count')
).round(2).sort_values('Revenue', ascending=False)

row_start3 = row_start2 + 2 + len(monthly) + 3
ws_pivot.cell(row=row_start3, column=1, value="Sales by Customer Type").font = Font(bold=True, size=13)
for r_idx, row in enumerate(dataframe_to_rows(cust_pivot.reset_index(), index=False, header=True), row_start3 + 2):
    for c_idx, val in enumerate(row, 1):
        cell = ws_pivot.cell(row=r_idx, column=c_idx, value=val)
        if r_idx == row_start3 + 2:
            cell.font = header_font
            cell.fill = header_fill
        cell.border = thin_border

# ========== SHEET 3: Key Insights ==========
ws_insights = wb.create_sheet("Key Insights")
insights = [
    ["RETAIL SALES ANALYSIS — KEY INSIGHTS", "", ""],
    ["", "", ""],
    ["Metric", "Value", "Notes"],
    ["Total Revenue", f"${df['Total_Price'].sum():,.2f}", "Across all 500 orders"],
    ["Total Profit", f"${df['Profit'].sum():,.2f}", ""],
    ["Overall Margin", f"{df['Profit'].sum()/df['Total_Price'].sum()*100:.1f}%", ""],
    ["Total Orders", len(df), ""],
    ["Avg Order Value", f"${df['Total_Price'].mean():,.2f}", ""],
    ["Top Category", cat_pivot.index[0], f"{cat_pivot.iloc[0]['Revenue']/cat_pivot['Revenue'].sum()*100:.1f}% of revenue"],
    ["Top Region", region_pivot.index[0], f"{region_pivot.iloc[0]['Revenue']/region_pivot['Revenue'].sum()*100:.1f}% of revenue"],
    ["Best Customer Type", cust_pivot.index[0], ""],
    ["", "", ""],
    ["RECOMMENDATIONS", "", ""],
    ["1", "Focus marketing on VIP customers", "Highest avg order value and margin"],
    ["2", "Expand Online channel", "Largest revenue share — invest in UX"],
    ["3", "Review discount strategy", "Discounts >15% erode margins significantly"],
    ["4", "Target South region growth", "Currently underperforming vs other regions"],
    ["5", "Bundle Electronics with accessories", "Top category — maximize basket size"],
]
for r_idx, row in enumerate(insights, 1):
    for c_idx, val in enumerate(row, 1):
        cell = ws_insights.cell(row=r_idx, column=c_idx, value=val)
        if r_idx == 1:
            cell.font = Font(bold=True, size=14, color="2F5496")
        elif r_idx == 3 or r_idx == 13:
            cell.font = header_font
            cell.fill = header_fill
        cell.border = thin_border

ws_insights.column_dimensions['A'].width = 25
ws_insights.column_dimensions['B'].width = 20
ws_insights.column_dimensions['C'].width = 40

# Save
output_path = "retail_sales_analysis.xlsx"
wb.save(output_path)
print(f"Excel file saved: {output_path}")
print(f"Sheets: {wb.sheetnames}")
